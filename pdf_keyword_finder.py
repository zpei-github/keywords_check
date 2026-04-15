"""
PDF关键字搜索工具
功能：
- 支持搜索多个关键字
- 返回关键字所在的完整句子
- 正确处理跨行句子
- 正确处理跨页句子
- 按重要性排序导出Excel

性能优化：
- 使用列表+join代替字符串拼接
- 合并关键字为单个正则模式，一次遍历
- 二分查找页码，O(log n)复杂度
- 预编译正则表达式
"""

import fitz  # PyMuPDF
import re
from typing import List, Dict, Tuple, Union
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont


# 预编译正则表达式
WHITESPACE_PATTERN = re.compile(r'\s+')
EXTRA_BOUNDARY_CHARS = set(',，;；')
BOUNDARY_CHARS = set('。！？!?')
IGNOR_CHARS = set('(（注')

# 默认的页眉页脚检测参数
DEFAULT_HEADER_RATIO = 0.15   # 页眉区域占比（顶部）
DEFAULT_FOOTER_RATIO = 0.85   # 页脚区域占比（底部）
DEFAULT_REPEAT_THRESHOLD = 0.3  # 重复率阈值
DEFAULT_MIN_TEXT_LENGTH = 3   # 最小文本长度


def detect_noise_blocks(
    all_blocks: List[Dict],
    keywords: List[str] = None,
    header_ratio: float = DEFAULT_HEADER_RATIO,
    footer_ratio: float = DEFAULT_FOOTER_RATIO,
    repeat_threshold: float = DEFAULT_REPEAT_THRESHOLD,
    min_text_length: int = DEFAULT_MIN_TEXT_LENGTH
) -> Tuple[set, List[Dict]]:
    """
    自动检测页眉页脚和水印block

    策略：
    1. 统计每个文本在不同页面出现的次数
    2. 结合位置（边缘区域）+ 重复率判断是否为噪声
    3. 关键字保护：包含关键字的block不当作噪声

    参数:
        all_blocks: 所有block信息
        keywords: 关键字列表（用于保护用户关心的内容）
        header_ratio: 页眉区域占比
        footer_ratio: 页脚区域占比
        repeat_threshold: 重复率阈值
        min_text_length: 最小文本长度

    返回:
        (需要过滤的block索引集合, 噪声信息列表)
    """
    if not all_blocks:
        return set(), []
    total_pages = len(set(b['page'] for b in all_blocks))
    if total_pages == 0:
        return set(), []
    # 构建关键字正则模式（用于保护包含关键字的内容）
    keyword_pattern = None
    if keywords:
        escaped_keywords = [re.escape(kw) for kw in keywords]
        keyword_pattern = re.compile('(' + '|'.join(escaped_keywords) + ')', re.IGNORECASE)
    else:
        return set(), []

    # 收集每个页面的高度信息
    page_heights = {}
    for block in all_blocks:
        page_num = block['page']
        if page_num not in page_heights:
            page_heights[page_num] = block.get('page_height', 792)  # 默认A4高度

    # 统计每个文本内容在不同页面出现的次数（用于检测重复的页眉页脚）
    text_occurrences = {}
    for block in all_blocks:
        text = block['text']
        if len(text) >= min_text_length:  # 只统计有意义的文本
            if text:
                if text not in text_occurrences:
                    text_occurrences[text] = set()
                text_occurrences[text].add(block['page'])

    # 计算每个文本的重复率
    repeated_texts = set()
    for text, pages in text_occurrences.items():
        repeat_rate = len(pages) / total_pages
        if repeat_rate >= repeat_threshold:
            repeated_texts.add(text)

    # 判断哪些block应该被过滤
    noise_indices = set()
    noise_info = []  # 记录被过滤的block信息

    for idx, block in enumerate(all_blocks):
        text = block['text']

        y0 = block.get('y0', 0)
        y1 = block.get('y1', 0)
        page_height = page_heights.get(block['page'], 792)

        # 规则1（重要）：关键字保护 - 包含关键字的block不当作噪声
        if keyword_pattern and keyword_pattern.search(text):
            continue  # 保留

        '''
        规则1.5（重要）：跨行关键字保护 - 边缘位置block拼接上一个/下一个block检测关键字
        
        当关键字被pdf文件中的换行符切分后, 正则表达式是无法匹配的, 需要重新拼接之后再匹配
        '''
        if idx == 0 and idx + 1 < len(all_blocks):
            next_block = all_blocks[idx + 1]
            next_text = next_block['text']
            
            combined = text + next_text
            if keyword_pattern and keyword_pattern.search(combined):
                continue
        elif idx + 1 < len(all_blocks) and idx - 1 >= 0:
            last_block = all_blocks[idx - 1]
            last_text = last_block['text']

            next_block = all_blocks[idx + 1]
            next_text = next_block['text']
            
            combined = last_text+ text + next_text
            if keyword_pattern and keyword_pattern.search(combined):
                continue
        else:
            last_block = all_blocks[idx - 1]
            last_text = last_block['text']
            combined = last_text+ text
            if keyword_pattern and keyword_pattern.search(combined):
                continue

        # 边缘判断和重复率判断基准
        is_at_edge = (y0 < page_height * header_ratio) or (y1 > page_height * footer_ratio)
        is_repeated = text in repeated_texts

        # 规则3：位置 + 重复率判断
        if is_at_edge and is_repeated:
            noise_indices.add(idx)
            noise_info.append({
                'page': block['page'],
                'block_no': block['block_no'],
                'text': text[:50] + '...' if len(text) > 50 else text,
                'repeat_rate': len(text_occurrences.get(text, set())) / total_pages,
                'reason': '边缘位置+高频重复',
                'position': '页眉' if y0 < page_height * header_ratio else '页脚'
            })

        # 规则2：包含数字的文本且在边缘位置（可能是页码、页眉页脚中的页码）
        if re.search(r'\d', text) and is_at_edge and len(text):
            noise_indices.add(idx)
            noise_info.append({
                'page': block['page'],
                'block_no': block['block_no'],
                'text': text,
                'repeat_rate': 0,
                'reason': '边缘位置+带有数字',
                'position': '页眉' if y0 < page_height * header_ratio else '页脚'
            })
            continue

    return noise_indices, noise_info


def get_page_text_with_layout(
    pdf_path: str,
    keywords: List[str] = None,
    auto_clean_noise: bool = False,
    check_pages: int = None,
    header_ratio: float = DEFAULT_HEADER_RATIO,
    footer_ratio: float = DEFAULT_FOOTER_RATIO,
    repeat_threshold: float = DEFAULT_REPEAT_THRESHOLD
) -> Tuple[str, List[Dict], List[Dict], List[int]]:
    """
    获取PDF文本，将每个block去除换行符后拼接

    优化：使用列表+join代替字符串拼接，时间复杂度从O(n²)降至O(n)

    参数:
        pdf_path: PDF文件路径
        keywords: 关键字列表（用于保护用户关心的内容，不过滤包含关键字的block）
        auto_clean_noise: 是否自动检测并去除页眉页脚页码
        header_ratio: 页眉区域占比（默认15%）
        footer_ratio: 页脚区域占比（默认85%）
        repeat_threshold: 重复率阈值（默认30%）
        check_pages: 采样检测的页数（默认全部）

    返回:
        (拼接后的完整文本, block信息列表, 噪声信息列表)
    """
    doc = fitz.open(pdf_path)

    # 第一遍：收集所有block信息（包括坐标）
    raw_blocks = []
    page_heights = {}
    
    # 前缀和记录每页字符数
    page_prifix_sum = []
    page_prifix_sum.append(0)

    for page_num in range(len(doc)):
        page = doc[page_num]
        page_height = page.rect.height
        page_heights[page_num + 1] = page_height
        blocks = page.get_text_blocks()

        # page_prifix_sum初始化
        page_prifix_sum.append(0)

        for block in blocks:
            x0, y0, x1, y1, text, block_no, block_type = block
            # 只处理文本块
            if block_type == 0 and text.strip():
                # 去除block内的换行符
                clean_text = text.replace('\n', '').replace('\r', '')
                clean_text = WHITESPACE_PATTERN.sub(' ', clean_text).strip()
                if clean_text:
                    raw_blocks.append({
                        'page': page_num + 1,
                        'block_no': block_no,
                        'text': clean_text,
                        'x0': x0, 'y0': y0, 'x1': x1, 'y1': y1,
                        'page_height': page_height
                    })
            

    doc.close()

    # 自动检测噪声block
    noise_indices = set()
    noise_info = []
    if auto_clean_noise:
        # 限制检测的页面数量以提高性能
        blocks_to_check = raw_blocks
        noise_indices, noise_info = detect_noise_blocks(
            blocks_to_check,
            keywords=keywords,
            header_ratio=header_ratio,
            footer_ratio=footer_ratio,
            repeat_threshold=repeat_threshold
        )

    # 第二遍：过滤噪声block并拼接
    text_parts = []
    block_info = []
    current_pos = 0


    for idx, block in enumerate(raw_blocks):
        # 跳过噪声block
        if idx in noise_indices:
            continue
        
        clean_text = block['text']
        block_info.append({
            'page': block['page'],
            'block_no': block['block_no'],
            'text': clean_text,
            'position': current_pos
        })
        text_parts.append(clean_text)
        current_pos += len(clean_text)
        # 记录清洗后的全文分布在每页中的字符数
        page_prifix_sum[block['page']] += len(clean_text)

    if(len(page_prifix_sum) > 1):
        for i in range(1 , len(page_prifix_sum)):
            page_prifix_sum[i] += page_prifix_sum[i - 1]
    full_text = ''.join(text_parts)  # O(n)拼接

    if auto_clean_noise:
        original_count = len(raw_blocks)
        filtered_count = len(block_info)
        print(f"过滤完成: {original_count} -> {filtered_count} 个block (去除 {original_count - filtered_count} 个噪声)")

    return full_text, block_info, noise_info, page_prifix_sum

def find_keywords_in_text(
    full_text: str,
    keywords: List[str],
    context_chars: int,
    front_window: int
) -> Tuple[List[Dict], List[Dict]]:
    """
    在文本中搜索关键字，返回完整句子（去重后）

    优化：合并所有关键字为单个正则模式，一次遍历完成匹配
    """
    if not keywords:
        return [],[]

    # 构建合并的正则模式：(keyword1|keyword2|...)
    # 使用 re.escape 确保特殊字符正确处理
    escaped_keywords = [re.escape(kw) for kw in keywords]
    pattern = re.compile('(' + '|'.join(escaped_keywords) + ')', re.IGNORECASE)

    # 一次遍历收集所有匹配
    all_matches = []
    for match in pattern.finditer(full_text):
        matched_text = match.group()
        # 找到匹配的是哪个关键字（保持原始大小写）
        for keyword in keywords:
            if keyword.lower() == matched_text.lower():
                all_matches.append({
                    'keyword': keyword,
                    'start': match.start(),
                    'end': match.end()
                })
                break

    if not all_matches:
        return [],[]

    # 提取句子并根据范围重叠进行合并
    sentence_list = []

    for match in all_matches:
        sentence, sentence_start, sentence_end = extract_sentence_from_text(
            text = full_text,
            start_pos = match['start'],
            end_pos = match['end'],
            context_chars = context_chars,
            front_window= front_window
        )

        sentence_list.append({
            'sentence': sentence,
            'keywords': {match['keyword']},
            'sentence_start': sentence_start,
            'sentence_end': sentence_end,
            'position': match['start']
        })

    # 按句子开始位置, 句子结束位置分别进行排序
    sentence_list.sort(key=lambda x: (x['sentence_start'], x['sentence_end']))

    # 合并重叠的句子
    merged = []
    for item in sentence_list:
        if not merged:
            merged.append(item)
            continue
        last = merged[-1]

        if item['sentence_end'] >= last['sentence_end'] and item['sentence_start'] == last['sentence_start']:
            item['keywords'].update(last['keywords'])

            # 取最靠前的关键字position
            item["position"] = min(item["position"], last["position"])
            merged[-1] = item
        else:
            merged.append(item)
    return merged, all_matches


def extract_sentence_from_text(text: str, start_pos: int, end_pos: int, context_chars: int, front_window:int) -> Tuple[str, int, int]:
    """从文本中提取包含关键字的完整句子"""
    text_len = len(text)

    # 向前找句子开始
    min_bound = max(0, start_pos - context_chars)  # 向前边界
    front_start = max(0, start_pos - front_window)
    sentence_start = front_start  # 默认从front_start开始

    for i in range(front_start, -1, -1):
        current_char = text[i]
        # 边界检查：下一字符是否存在
        next_char = text[i + 1] if i + 1 < text_len else ''

        # 在 context_chars 范围内：只检测 BOUNDARY_CHARS
        # 突破边界后：同时检测 BOUNDARY_CHARS 和 EXTRA_BOUNDARY_CHARS
        if i >= min_bound:
            # 范围内：遇到句末标点且下一字符不在 IGNOR_CHARS 中才截断
            if current_char in BOUNDARY_CHARS and next_char not in IGNOR_CHARS:
                sentence_start = i + 1
                break
        else:
            # 突破边界：遇到任何分界符都截断
            if current_char in BOUNDARY_CHARS or current_char in EXTRA_BOUNDARY_CHARS:
                sentence_start = i + 1
                break

    # 向后找句子结束
    max_bound = min(text_len, end_pos + context_chars)  # 向后边界
    sentence_end = text_len  # 默认到文本末尾

    for i in range(end_pos, text_len):
        current_char = text[i]

        # 在 context_chars 范围内：只检测 BOUNDARY_CHARS
        # 突破边界后：同时检测 BOUNDARY_CHARS 和 EXTRA_BOUNDARY_CHARS
        if i < max_bound:
            # 范围内：遇到句末标点就截断
            if current_char in BOUNDARY_CHARS:
                sentence_end = i + 1
                break
        else:
            # 突破边界：遇到任何分界符都截断
            if current_char in BOUNDARY_CHARS or current_char in EXTRA_BOUNDARY_CHARS:
                sentence_end = i + 1
                break

    sentence = text[sentence_start:sentence_end].strip()
    # 清理多余空格
    sentence = WHITESPACE_PATTERN.sub(' ', sentence)
    return sentence, sentence_start, sentence_end


def find_keywords_in_pdf(
    pdf_path: str,
    context_rich: int,
    front_window: int,
    keywords: List[str] | Dict[str, int],
    output_file: str | None = None,
    excel_file: str | None = None,
    auto_clean_noise: bool = False,
    header_ratio: float = DEFAULT_HEADER_RATIO,
    footer_ratio: float = DEFAULT_FOOTER_RATIO,
    repeat_threshold: float = DEFAULT_REPEAT_THRESHOLD
) -> Dict:
    """
    在PDF中搜索关键字并返回结果

    参数:
        pdf_path: PDF文件路径
        keywords: 关键字列表或字典（关键字->分数）
        output_file: 可选的输出文件路径（txt）
        excel_file: 可选的Excel输出文件路径（按重要性排序）
        auto_clean_noise: 是否自动检测并去除页眉页脚页码
        header_ratio: 页眉区域占比（默认15%）
        footer_ratio: 页脚区域占比（默认85%）
        repeat_threshold: 重复率阈值（默认30%）
    """
    print(f"正在打开PDF文件: {pdf_path}")

    # 支持字典格式（关键字+分数）
    if isinstance(keywords, dict):
        keywords_point = keywords
        keywords_list = list(keywords.keys())
    else:
        keywords_list = keywords
        # 默认每字1分
        keywords_point = {k: 1 for k in keywords_list}

    # 获取拼接后的文本（传入keywords用于保护用户关心的内容）
    full_text, block_info, noise_info, page_prifix_sum = get_page_text_with_layout(
        pdf_path,
        keywords=keywords_list,
        auto_clean_noise=auto_clean_noise,
        check_pages=None,
        header_ratio=header_ratio,
        footer_ratio=footer_ratio,
        repeat_threshold=repeat_threshold
    )
    print(f"文本总长度: {len(full_text)} 字符, 共 {len(block_info)} 个文本块")

    # 搜索关键字
    print(f"正在搜索关键字: {keywords_list}")
    results, all_matchs = find_keywords_in_text(full_text, keywords_list, context_rich, front_window)

    print(f"\n找到 {len(results)} 处匹配")

    # 计算每个句子的总分
    for result in results:
        score = sum(keywords_point.get(kw, 1) for kw in result['keywords'])
        result['score'] = score

    # 按页码组织结果
    page_results = {}

    # results已按sentence_start升序排列，使用双指针优化
    last_page = 1
    last_end_page = 1
    for result in results:
        # 使用 page_prifix_sum 前缀和计算页码（更准确）
        pos = result.get('sentence_start', result['position'])
        sentence_end = result.get('sentence_end', result['position'])

        # 从上次位置继续向下遍历（利用升序排列）
        while last_page < len(page_prifix_sum) and pos >= page_prifix_sum[last_page]:
            last_page += 1
        page_num = last_page

        while last_end_page < len(page_prifix_sum) and sentence_end >= page_prifix_sum[last_end_page]:
            last_end_page += 1
        end_page = last_end_page

        # 如果起始页和结束页不同，标记为跨页
        is_cross_page = (end_page != page_num)
        result['page'] = page_num
        result['is_cross_page'] = is_cross_page
        result['end_page'] = end_page

        if page_num not in page_results:
            page_results[page_num] = []
        page_results[page_num].append(result)

    # 保存到txt文件
    if output_file:
        export_to_txt(
            output_file=output_file,
            pdf_path=pdf_path,
            all_matchs=all_matchs,
            keywords_list=keywords_list,
            keywords_point=keywords_point,
            results=results,
            page_results=page_results,
            noise_info=noise_info
        )

    # 导出Excel（按重要性排序）
    if excel_file:
        export_to_excel(results, excel_file, pdf_path, keywords_point)
        print(f"Excel已保存到: {excel_file}")

    return {
        'total_matches': len(results),
        'by_page': page_results,
        'all_results': results,
        'noise_info': noise_info
    }


def export_to_txt( 
		output_file: str,
		pdf_path: str,
		keywords_list: List[str],
		keywords_point: Dict[str, int],
		all_matchs: List[Dict],
		results: List[Dict],
		page_results: Dict,
		noise_info: List[Dict]
 ):
		"""
		将关键字搜索结果、统计信息及噪音检测结果导出到txt文件
		参数:
			output_file: 输出txt文件路径
			pdf_path: PDF文件路径
			keywords_list: 搜索关键字列表
			keywords_point: 关键字分数字典
			all_matchs: 所有关键字匹配点信息
			results: 所有搜索结果列表
			page_results: 按页码组织的搜索结果字典
			noise_info: 噪音检测信息列表
		"""
          
		with open(output_file, 'w', encoding='utf-8') as f:
			# 1. 基础信息
			f.write("PDF关键字搜索结果及深度分析报告\n")
			f.write("=" * 80 + "\n")
			f.write(f"PDF文件: {pdf_path}\n")
			keywords_info = ', '.join([f"{k}({v}分)" for k, v in keywords_point.items()])
			f.write(f"搜索关键字: {keywords_info}\n")
			# 2. 统计结果
			f.write("\n" + "=" * 80 + "\n")
			f.write("【一、 统计结果】\n")
			f.write("-" * 80 + "\n")
			f.write(f"总匹配句子数: {len(results)}\n")
			total_score = sum(r.get('score', 0) for r in results)
			f.write(f"总重要性得分: {total_score}\n")
			# 按关键字统计匹配次数及贡献分数
			keyword_counts = {}
			keyword_scores = {}
			for match in all_matchs:
				kw = match['keyword']
				keyword_counts[kw] = keyword_counts.get(kw, 0) + 1
				keyword_scores[kw] = keyword_scores.get(kw, 0) + keywords_point.get(kw, 1)
			f.write("\n各关键字命中详情:\n")
			for kw in keywords_list:
				count = keyword_counts.get(kw, 0)
				if count > 0:
					score = keyword_scores.get(kw, 0)
					f.write(f"  - {kw}: 命中 {count} 次, 贡献得分 {score}\n")
			# 跨页统计
			cross_page_count = sum(1 for r in results if r.get('is_cross_page', False))
			if cross_page_count > 0:
				f.write(f"\n跨页句子数: {cross_page_count}\n")                         

			# 4. 搜索结果详情（按页码顺序）
			f.write("\n" + "=" * 80 + "\n")
			f.write("【三、 搜索结果详情 - 按页码顺序】\n")
			f.write("-" * 80 + "\n")
			for page_num in sorted(page_results.keys()):
				f.write(f"\n--- 第 {page_num} 页 ---\n")
				for idx, r in enumerate(page_results[page_num], 1):
					keywords_str = ', '.join(r['keywords'])
					is_cross_page = r.get('is_cross_page', False)
					end_page = r.get('end_page', page_num)
					page_info = f"(跨页至第{end_page}页)" if is_cross_page else ""
					# 文本高亮替换
					sentence = r['sentence']
					f.write(f"\n  [{idx}] (得分:{r.get('score',0)}) 关键字: {keywords_str} {page_info}\n")
					f.write(f"   {sentence}\n")
                         

			# 5. 噪音检测结果
			f.write("\n" + "=" * 80 + "\n")
			f.write("【四、 噪音检测结果】\n")
			f.write("-" * 80 + "\n")
			if noise_info:
				f.write(f"共检测到 {len(noise_info)} 个可能被过滤的页眉/页脚/水印block:\n\n")
				unique_noise = {}
				for item in noise_info:
					key = item['text']
					if key not in unique_noise:
						unique_noise[key] = item
				# 按位置分类展示，更加清晰
				headers = [item for item in unique_noise.values() if item['position'] == '页眉']
				footers = [item for item in unique_noise.values() if item['position'] == '页脚']
				if headers:
					f.write(">>页眉噪音:\n")
					for item in headers:
						f.write(f"    - 重复率: {item['repeat_rate']:.1%} | 原因: {item['reason']}\n")
						f.write(f"      内容: {item['text']}\n")
				if footers:
					f.write("\n >>页脚噪音:\n")
					for item in footers:
						f.write(f"    - 重复率: {item['repeat_rate']:.1%} | 原因: {item['reason']}\n")
						f.write(f"      内容: {item['text']}\n")
			else:
				f.write("未检测到明显的页眉/页脚/水印噪音。\n")
		print(f"\n深度分析报告已保存到: {output_file}")


def export_to_excel(results: List[Dict], excel_file: str, pdf_path: str, keywords_point: Dict[str, int]):
    sorted_results = sorted(results, key=lambda x: x.get('score', 0), reverse=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "关键字搜索结果"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=12, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    wrap_alignment = Alignment(wrap_text=True, vertical='top')

    # 3.1.5 正确格式：InlineFont
    red_font = InlineFont(color='00FF0000')
    default_font = InlineFont(color='00000000')

    ws.merge_cells('A1:H1')
    ws['A1'] = f"PDF关键字搜索结果 - {pdf_path}"
    ws['A1'].font = Font(bold=True, size=14)

    ws.merge_cells('A2:H2')
    keywords_info = ', '.join([f"{k}({v}分)" for k, v in keywords_point.items()])
    ws['A2'] = f"搜索关键字: {keywords_info}"
    ws['A2'].font = Font(italic=True)

    headers = ['排名', '重要性', '得分', '始页', '终页', '包含关键字', '完整句子', '备注']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 60
    ws.column_dimensions['H'].width = 15

    max_score = max((r.get('score', 0) for r in sorted_results), default=1)

    for idx, result in enumerate(sorted_results, 1):
        row = idx + 4
        score = result.get('score', 0)

        if max_score > 0:
            ratio = score / max_score
            if ratio >= 0.8:
                importance = "★★★★★"
                importance_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
            elif ratio >= 0.6:
                importance = "★★★★"
                importance_fill = PatternFill(start_color="FFA94D", end_color="FFA94D", fill_type="solid")
            elif ratio >= 0.4:
                importance = "★★★"
                importance_fill = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
            elif ratio >= 0.2:
                importance = "★★"
                importance_fill = PatternFill(start_color="ADE25D", end_color="ADE25D", fill_type="solid")
            else:
                importance = "★"
                importance_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        else:
            importance = "★"
            importance_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        ws.cell(row=row, column=1, value=idx).border = thin_border
        ws.cell(row=row, column=2, value=importance).border = thin_border
        ws.cell(row=row, column=2).fill = importance_fill
        ws.cell(row=row, column=3, value=score).border = thin_border

        # 始页和终页
        start_page = result.get('page', result.get('page', 1))
        end_page = result.get('end_page', start_page)
        ws.cell(row=row, column=4, value=start_page).border = thin_border
        ws.cell(row=row, column=5, value=end_page).border = thin_border

        # 包含关键字
        ws.cell(row=row, column=6, value=', '.join(result['keywords'])).border = thin_border
        ws.cell(row=row, column=6).alignment = wrap_alignment

        # ===================== 标红逻辑 =====================
        sentence = result['sentence']
        keywords = result.get('keywords', [])
        sentence_cell = ws.cell(row=row, column=7)

        if not keywords:
            sentence_cell.value = sentence
        else:
            parts = []
            current = 0
            pattern = '|'.join(re.escape(k) for k in keywords)
            for match in re.finditer(pattern, sentence):
                s, e = match.span()
                if s > current:
                    parts.append(TextBlock(default_font, sentence[current:s]))
                parts.append(TextBlock(red_font, sentence[s:e]))
                current = e
            if current < len(sentence):
                parts.append(TextBlock(default_font, sentence[current:]))

            sentence_cell.value = CellRichText(parts)
        # ========================================================

        sentence_cell.border = thin_border
        sentence_cell.alignment = wrap_alignment

        # 备注栏：显示跨页信息
        is_cross_page = result.get('is_cross_page', False)
        remark = "跨页" if is_cross_page else ""
        remark_cell = ws.cell(row=row, column=8, value=remark)
        remark_cell.border = thin_border
        if is_cross_page:
            remark_cell.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")

        line_count = max(1, len(sentence) // 60 + 1)
        ws.row_dimensions[row].height = min(60, line_count * 15)

    ws.freeze_panes = 'A5'
    wb.save(excel_file)


# 使用示例
if __name__ == "__main__":
    # 示例：搜索单个PDF文件
    pdf_path = r"E:\Desktop\招标文件-副本.pdf"  # 替换为你的PDF文件路径

    # 定义要搜索的关键字及分数
    keywords_point = {
            "提供": 4,
            "提交": 4,
            "递交": 4,
            "出具": 4,
            "响应": 4,
            "加盖": 5,
            "承诺": 9,
            "授权": 6,
            "证明": 9,
            "公章": 7,
            "鲜章": 7,
            "报告": 5,
            "签字": 3,
            "说明": 3,
            "证书": 4,
            "盖单位章": 7,
            "盖章": 7,
            "签章": 7,
            "法人章": 7,
            "必须": 4
    }

    # 执行搜索（直接传字典，会自动计算分数）
    results = find_keywords_in_pdf(
        pdf_path=pdf_path,
        keywords=keywords_point,
        context_rich=100,
        front_window= 0,
        output_file=r"E:\Desktop\output.txt",  # 可选：保存txt结果
        excel_file=r"E:\Desktop\output.xlsx",  # 可选：保存Excel结果（按重要性排序）

        #==================噪声检测配置==================
        auto_clean_noise=True,  # 开启自动检测页眉页脚和水印
        header_ratio=0.05,      # 页眉区域占比
        footer_ratio=0.95,      # 页脚区域占比
        repeat_threshold=0.8    # 重复率阈值
    )
